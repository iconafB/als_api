from fastapi import APIRouter,HTTPException,Depends,UploadFile,File,Query,Request,status
from sqlmodel import Session,select,update,delete
from sqlalchemy import func,text

from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.dialects import postgresql
from typing import Annotated,List,Dict,Any
import openpyxl
import random
from io import BytesIO
import re 
import os
import pandas as pd
import string
import time
from sqlalchemy import insert,and_,or_
from sqlalchemy.sql import func
from datetime import datetime
from models.campaigns import Campaign_Dedupe,Deduped_Campaigns,dedupe_campaigns_tbl
from models.leads import info_tbl
from schemas.dedupe_campaigns import CreateDedupeCampaign,SubmitDedupeReturnSchema,ManualDedupeListReturn,StatusData,EnrichedData,CreateDedupeCampaign,DeleteCamapignSchema,UpdateDedupeCampaign
from database.database import get_session
from utils.auth import get_current_user,get_current_active_user
from utils.status_data import get_status_tuple,insert_data_into_finance_table,insert_data_into_location_table,insert_data_into_contact_table,insert_data_into_employment_table,insert_data_into_car_table
from database.database import engine

from utils.logger import define_logger
from schemas.dedupes import DataInsertionSchema
from schemas.status_data_routes import InsertStatusDataResponse,InsertEnrichedDataResponse

dedupe_logger=define_logger("als dedupe campaign logs","logs/dedupe_route.log")

status_data_logger=define_logger("als status logger logs","logs/status_data.log")

dedupe_routes=APIRouter(tags=["Dedupes"],prefix="/dedupes")
#dedupe campaign crud
@dedupe_routes.post("",status_code=status.HTTP_201_CREATED,description="create dedupe campaign by provide the campaign name,campaign code and branch",response_model=dedupe_campaigns_tbl)

async def create_dedupe_campaign(dedupe:CreateDedupeCampaign,session:Session=Depends(get_session),user=Depends(get_current_active_user)):
    try:
        dedupe_campaign_query=select(dedupe_campaigns_tbl).where((dedupe_campaigns_tbl.camp_code==dedupe.camp_code)&(dedupe_campaigns_tbl.campaign_name==dedupe.campaign_name)&(dedupe_campaigns_tbl.branch==dedupe.branch))
        dedupe_campaign=session.exec(dedupe_campaign_query).first()

        if not dedupe_campaign:
            dedupe_logger.info(f"user:{user.id} with email:{user.email} attempted to create:{dedupe.campaign_name} with campaign code:{dedupe.camp_code} at branch:{dedupe.branch}")
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,detail=f"dedupe campaign:{dedupe.campaign_name} already exist")
        payload=dedupe.model_dump()
        campaign=dedupe_campaigns_tbl.model_validate(payload)
        session.add(campaign)
        session.commit()
        session.refresh(campaign)
        dedupe_logger.info(f"user:{user.id} with email:{user.email} created campaign:{dedupe.campaign_name} with campaign code:{dedupe.camp_code} at branch:{dedupe.branch}")
        return campaign
    
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        session.rollback()
        return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"error occurred while creating dedupe campaign:{dedupe.campaign_name} with campaign code:{dedupe.camp_code} for branch:{dedupe.branch}")



@dedupe_routes.get("/{camp_code}",description="Get dedupe campaign by providing the campaign code",status_code=status.HTTP_200_OK,response_model=dedupe_campaigns_tbl)

async def get_dedupe_campaign(camp_code:str,session:Session=Depends(get_session),user=Depends(get_session)):
    try:
        #find the campaign using campaign code
        campaign_query=select(dedupe_campaigns_tbl).where(dedupe_campaigns_tbl.camp_code==camp_code)
        
        campaign=session.exec(campaign_query).first()

        if not campaign:
            dedupe_logger.info(f"user:{user.id} with email:{user.email} requested campaign:{camp_code} and it does not exist")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"campaign:{camp_code} does not exist")
        
        return campaign
    
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occuurred while retrieving campaign:{camp_code}")


#get all dedupe campaigns

@dedupe_routes.get("",description="Get all dedupe camapigns",status_code=status.HTTP_200_OK,response_model=List[dedupe_campaigns_tbl])

async def get_all_dedupe_campaigns(session:Session=Depends(get_session),user=Depends(get_current_user)):
    try:
        dedupe_logger.info(f"user:{user.id} with email:{user.email} retrieved all the dedupe campaigns")
        all_campaigns_query=select(dedupe_campaigns_tbl)
        all_dedupe_campaign=session.exec(all_campaigns_query).all()
        
        return all_dedupe_campaign
    
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred while fetching dedupe campaigns")

#get active dedupe campaigns
@dedupe_routes.get("/active",status_code=status.HTTP_200_OK,description="Get all active dedupe campaigns",response_model=List[dedupe_campaigns_tbl])

async def get_active_dedupe_campaigns(session:Session=Depends(get_session),user=Depends(get_current_active_user)):
    try:
        active_dedupe_campaign_queries=select(dedupe_campaigns_tbl).where(dedupe_campaigns_tbl.is_active==True)
        active_dedupe_campaigns=session.exec(active_dedupe_campaign_queries).all()

        dedupe_logger.info(f"user:{user.id} with email:{user.email} retrieved {len(active_dedupe_campaigns)} dedupe campaigns")

        return active_dedupe_campaigns
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while fetching all active campaigns")

#deactivate/delete campaigns
@dedupe_routes.patch("/delete/{camp_code}",status_code=status.HTTP_202_ACCEPTED)

async def delete_dedupe_campaign(camp_code:str,session:Session=Depends(get_session),user=Depends(get_current_active_user)):
    try:
        campaign_query=select(dedupe_campaigns_tbl).where((dedupe_campaigns_tbl.camp_code==camp_code)&(dedupe_campaigns_tbl.is_active==True))
        campaign=session.exec(campaign_query).first()
        if not campaign:
            dedupe_logger.info(f"campaign with campaign code:{camp_code} does not exist")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"camapign with campaign code:{camp_code} does not exist")
        campaign.is_active=False
        session.add(campaign)
        session.commit()
        dedupe_logger.info(f"user:{user.id} with email:{user.email} deleted dedupe campaign with code:{camp_code}")
        return DeleteCamapignSchema(campaign_code=camp_code,message=f"Dedupe campaign:{camp_code} deleted")
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while deleting dedupe campaign:{camp_code}")

#update campaign names,codes and branch should only be done by the db administrator

@dedupe_routes.patch("/update/{camp_code}",status_code=status.HTTP_200_OK,description="Update campaign name,campaign code,or branch only provide the field you want to update",response_model=dedupe_campaigns_tbl)

async def update_campaign(campaign:UpdateDedupeCampaign,session:Session=Depends(get_session),user=Depends(get_current_active_user)):
    try:
        campaign_query=select(dedupe_campaigns_tbl).where(dedupe_campaigns_tbl.camp_code==campaign.camp_code)
        campaign_found=session.exec(campaign_query).first()
        if not campaign_found:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"campaign:{campaign.camp_code}")
        update_campaign=campaign.dict(exclude_unset=True)
        for key,value in update_campaign.items():
            setattr(campaign_found,key,value)
        session.commit()
        session.refresh(campaign_found)
        dedupe_logger.info(f"dedupe campaign:{campaign.camp_code} with campaign name:{campaign.camp_name} updated by user:{user.id} with email:{user.email}")
        return campaign_found
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while updating the campaign:{campaign.camp_code}")

#load dedupe campaign

@dedupe_routes.post("")

async def load_dedupe_campaign():
    try:
        return True
    except Exception as e:
        dedupe_logger.error(f"{str(e)}")
        return HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"error occurred while loading the campaign")

@dedupe_routes.post("/add-dedupes-manually",status_code=status.HTTP_200_OK)

async def add_dedupes_manually(campaign_name:str=Query(description="Please provide the campaign name"),file:UploadFile=File(...,description="Please provide a file for a manual dedupe"),session:Session=Depends(get_session),user=Depends(get_current_user)):
    
    #we are going to use these rows
    all_rows=[]

    file_path=f"temp_{file.filename}"

    #check the extension of the file format being uploaded
    if not file.filename.endswith((".csv",".xlsx",".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,detail="Invalid file format")
    
    try:
        #read the file's content into a BytesIO object(memory file)
        file_content=await file.read()
        file_stream=BytesIO(file_content)

        if file.filename.endswith(".csv"):
            #csv processing logic
            pass
        else:
            #check the campaign codes chiefs??
            workbook=openpyxl.load_workbook(file_stream,read_only=True)

            for sheet in workbook.worksheets:
                extracted_data=[
                  [
                      str(row[0] if row[0] is not None else ""),
                      str(row[1] if row[1] is not None else "")
                  ]
                  for row in sheet.iter_rows(min_col=1,max_col=2,values_only=True)
                  if row[0] is not None or row[1] is not None
                ]
                
                all_rows.extend(extracted_data)
                
            #insert the data into the campaign_dedupe using bulk insert, we still need the campaign name and is_verified field filled
            data_to_insert=[
                {
                    "id":row[0],
                    "cell":row[1],
                    "campaign_codes":campaign_name,
                    "is_verified":True
                }
                for row in all_rows
            ]
            session.bulk_insert_mappings(Campaign_Dedupe,data_to_insert)
            session.commit()
            session.close()

            return {"message":f"file with name:{file.filename} for campaign:{campaign_name} uploaded successfully"}

    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"{str(e)}")
    
    finally:
        #close the session
        session.close()
    

# this should an automated script that polls the dmasa api and post it to vc dial
@dedupe_routes.post("/submit-dedupe-return")

async def submit_dedupe_return(user=Depends(get_current_user)):

    return {"message":"submit dedupe return"}

#get the rowcount to get the number leads


#add dedupe list for a campaign, campaign code is an arguement and return a filename with leads

@dedupe_routes.post("/add-dedupe-list")

async def add_dedupe_list(camp_code:str,session:Session=Depends(get_session),user=Depends(get_current_user)):
   
    try:
        #fetch a dedupe campaign that matches the given campaign code
        campaign=session.exec(select(Deduped_Campaigns).where(Deduped_Campaigns.camp_code==camp_code)).first()
        #raise an exception if the dedupe campaign does not exist prompting the user to create a rule for that campaign
        
        if campaign==None:
            dedupe_logger.info(f"dedupe campaign with code:{camp_code} does not exist or it is not a dedupe campaign")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"Campaign with campaign code:{camp_code} does not exist or it is not a dedupe campaign")
        
        #extract the rule 
        if campaign.camp_rule==None:
            dedupe_logger.info(f"campaign rule for campaign:{campaign.camp_name} does not exist")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"campaign rule for campaign:{campaign.camp_name} does not exist")
        
        #fetch leads id and cell number from the information table
        #construct a dynamic sql query
        select_query=select(info_tbl.id,info_tbl.cell)

        if campaign.camp_rule.gender is not None:

            select_query=select_query.where(info_tbl.gender==campaign.camp_rule.gender)

        if campaign.camp_rule.derived_income is not None:
            select_query=select_query.where(info_tbl.derived_income>=campaign.camp_rule.derived_income)

        if campaign.camp_rule.minimum_salary is not None:
            select_query=select_query.where(info_tbl.salary>=campaign.camp_rule.minimum_salary)
        
        if campaign.camp_rule.maximum_salary is not None:
            select_query=select_query.where(info_tbl.salary<=campaign.camp_rule.maximum_salary)
        
        #consider date for the following query
        #execute the query and fetch the users with
        fetched_leads=session.exec(select_query).fetchall()

        if fetched_leads==None:
            dedupe_logger.info(f"zero leads match campaign:{campaign.camp_name} with code:{campaign.camp_code}")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"zero leads that matches this spec")
        
        #calculate the total leads
        if fetched_leads.count()>0:

            todaysdate=datetime.today().strftime('%Y-%m-%d')
            filename=camp_code + '-'+ todaysdate
            #insert leads inside the campaign_dedupe table
            suffix=''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(4))
            key=filename + suffix
        
        #print the key
        print(f"print the key:{key}")

        #new update data list with data for the campaign dedupes table
        updated_data=[(data[0],data[1],'P',key) for data in fetched_leads]

        campaign_object=[
            Campaign_Dedupe(id=data[0],cell=data[1],campaign_name=camp_code,status=data[2],code=data[3]) for data in updated_data
        ]  


        session_campaign_object=session.add(campaign_object)

        #questionable task

        if session_campaign_object == None:
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"error occurred while creating session object")
        
        #commit the session
        session.commit()
        dedupe_logger.info("updated the campaign dedupes table")
        #update data from the information table
        informatable_data=[(data[1],'DEDUPE')for data in fetched_leads]

        #Insert data into the information table
        information_object=[info_tbl(cell=data[0],extra_info=data[1]) for data in informatable_data]
        
        #Bulk insert into the information table
        session.add(information_object)
        #commit the data on the information table
        session.commit()

        dedupe_logger.info("updated information table")

        #search the filename and check if the file name has a slash(/) on its name
        if '/' in filename:
            new_filename=filename.replace("/","-")

        else:
            new_filename=filename
        
        #mistery
        df=pd.DataFrame(fetched_leads)
        df.to_excel(new_filename+'.xlsx',index=False)

        return {
            'Success':True,
            'filename':filename,
            'Key':key
        }
    
    except Exception as e:
        dedupe_logger.critical(f"crititcal error occurred:{e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")
    
#create a dedupe campaign

@dedupe_routes.post("/dedupe-campaigns",status_code=status.HTTP_200_OK)
async def create_dedupe_campaign(req:Request,campaign:CreateDedupeCampaign,session:Session=Depends(get_session),user=Depends(get_current_user)):
    
    try:
        #print the request object
        print("print the request object first")
        print(req)
        #search if a dedupe campaign exist

        deduped_campaign=session.exec(select(Deduped_Campaigns).where(Deduped_Campaigns.camp_name==campaign.campaign_name and Deduped_Campaigns.camp_code==campaign.campaign_code)).first()

        if deduped_campaign==None:
            dedupe_logger.info(f"campaign:{campaign.campaign_name} with campaign code:{campaign.campaign_code} already exist")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"campaign:{campaign.campaign_name} with campaign code:{campaign.campaign_code} already exist")
        
        #validation
        new_dedupe_campaign=Deduped_Campaigns(brach=campaign.branch,camp_name=campaign.campaign_name,camp_code=campaign.campaign_code,camp_rule={"minimum_salary":campaign.minimum_salary,"maximum_salary":campaign.maximum_salary,"derived_income":campaign.derived_income,"gender":campaign.gender,"limit":campaign.limit})
        if not new_dedupe_campaign:
            dedupe_logger.critical(f"error creating dedupe campaign:{campaign.campaign_name}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"error occurred while creating deduped campaign:{campaign.campaign_name}")
        session.add(new_dedupe_campaign)
        session.commit()
        session.refresh(new_dedupe_campaign)
        dedupe_logger(f"dedupe campaign:{campaign.campaign_name} successfully created")
        return new_dedupe_campaign
    
    except Exception as e:
        dedupe_logger.critical(f"error:{e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred while creating a dedupe campaign")




#submit dedupe return route

@dedupe_routes.post("/submit-dedupe-return",status_code=status.HTTP_200_OK)

async def submit_dedupe_return(data:SubmitDedupeReturnSchema,dedupe_file:UploadFile=File(...),session:Session=Depends(get_session)):
    try:
        #read the uploaded file baba
        file_contents=await dedupe_file.read()
        data_frame=pd.DataFrame(file_contents.splitlines())
        list_contents=data_frame.values.tolist()
        #dedupe list content

        dedupe_list=[]

        dedupe_list2=[str(d[0],'UTF-8') for d in list_contents]

        new_tuple_list=[d for d in dedupe_list2 if re.match('^\d{13}$',d)]

        for list in list_contents:
            #new tuple
            new_tuple=str(list[0],'UTF-8')
            if re.match('^\d{13}$',new_tuple):
                dedupe_list.append(new_tuple)

        #search for dedupe campaign that matches the campaign code
        dedupe_query=select(Campaign_Dedupe.code).where(Campaign_Dedupe.code==data.code)

        exec_dedupe_query=session.exec(dedupe_query).all()
        #raise an exception if the query fails

        if exec_dedupe_query==None:
            dedupe_logger.info(f"campaign dedupe:{data.campaign_name} with code:{data.code} does not exist")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"campaign dedupe:{data.campaign_name} with code:{data.code}")
        
        #count the number of campaigns returned
        campaign_count_query=(select(func.count()).select_from(Campaign_Dedupe).where(Campaign_Dedupe.code==data.code))
        #this value can also be zero,must be zero

        campaign_count=session.exec(campaign_count_query).scalar_one()

        #search the dedupe campaign
        dedupe_campaign_stmt=select(Deduped_Campaigns).where(Deduped_Campaigns.camp_code==data.campaign_code)
        
        dedupe_campaign=session.exec(dedupe_campaign_stmt).first()

        camp_found=False

        if dedupe_campaign==None:

            camp_found=False
            dedupe_logger.info(f"campaign:{data.campaign_name} with campaign code:{data.campaign_code} does not exist or is not a dedupe campaign")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"dedupe campaign:{data.campaign_name} with campaign code:{data.campaign_code} does not exist")
        
        else:
            camp_found=True
        
        #test validation
        valid=False
        validation_message=None

        dedupe_message=None
        if camp_found == True and len(dedupe_list)>0 and campaign_count >0:
            valid=True
            validation_message="Validation Passed"
            dedupe_message="Dedupe Campaign Exist"

        else:
            validation_message="Validation Failed"
            if not camp_found:
                dedupe_message="Campaign Entered is not a deduped campaign"
            if len(dedupe_list)==0:
                list_message="Empty list entered"
        
        #dedupe list tuple
        database_list=tuple(dedupe_list)

        #update statement query, questinable
        update_statement=(update(Campaign_Dedupe).values(status='R').where(Campaign_Dedupe.code==data.code,Campaign_Dedupe.id.in_(database_list)))
        #execute the query
        session.exec(update_statement)
        #commit the session
        session.commit()
        #poll the session object for any error in committing the object
        #select id from campaign_dedupe where status is 'P' and code matches the supplied 
        
        id_query_from_campaign_dedupe=select(Campaign_Dedupe.id).where(Campaign_Dedupe.code==data.code and Campaign_Dedupe.status=='P')
        
        id_campaign_dedupe_query_exec=session.exec(id_query_from_campaign_dedupe).all()
        #count the number of ids fetched

        if not id_campaign_dedupe_query_exec:
            dedupe_logger.info("ids not found")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"ids not found")
        #load ids into a list

        id_numbers=[n[0] for n in id_campaign_dedupe_query_exec if n[0] is not None]
        #place the database list in a tuple
        id_database_list=tuple(id_numbers)
        id_database_list_count=session.exec(id_query_from_campaign_dedupe).scalar_one()
        #if the count is greater than zero

        if id_database_list_count > 0:
            #delete statement,questionable in_ operator
            delete_statement=delete(Campaign_Dedupe).where(Campaign_Dedupe.id.in_(id_database_list))
            #update sql statement
            update_statement=update(info_tbl).where(info_tbl.id.in_(id_database_list)).values(extra_info=None)
            #delete session
            session.exec(delete_statement)
            session.commit()
            #update session on the information table
            session.exec(update_statement)
            session.commit()


        #delete statement on campaign dedupes where code = 'U'

        delete_statement_code_u=delete(Campaign_Dedupe).where(Campaign_Dedupe.code=='U')

        session.exec(delete_statement_code_u)
        session.commit()
        #poll the rows affected an raise an exception 
        
        return {
            "Success":True
        }
    


    
    except Exception as e:
        dedupe_logger.critical(f"{e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error:{e}")
    
#insert status data

@dedupe_routes.post("/insert_status_data",status_code=status.HTTP_200_OK,response_model=InsertStatusDataResponse)
#return the time taken for the queries, number of leads affected

async def insert_status_data(filename:str=Query(...,description="Provided the name of the filename with status data"),session:Session=Depends(get_session)):
    try:
        delta_time_1=time.time()
        status_data=[]
        #read a csv file that already exist on the server using the query parameter filename
        #check if the file exists
        if not os.path.exists(filename):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,detail=f"filename:{filename} does not exist on this system")
        
        #read the csv file into the pandas dataframe
        status_dataframe=pd.read_csv(filename)
        # convert the csv data into a list
        csv_list=status_dataframe.values.tolist()
        #append the uploaded list into the status data array
        status_data=status_data + csv_list
        #read the total number of leads
        leads_total=len(status_data)
        #seconds passed
        delta_time2=time.time()
        #total time difference
        total_time=delta_time2 - delta_time_1
        #rows list 
        rows=[]
        status_data=['0'+ str(row[15]) for row in status_data]

        for row in status_data:
            
            cell='0'+str(row[15])

            if re.match('^(\d{10})?$', str(cell)):

                if row[1] is not None:
                    date_created_old=row[1].split('')[0]
                #test the id number if its has 13 characters
                idnum=str(row[3])
                salary=str(row[2])
                name = row[6]
                surname = row[7]
                address1 = row[9]
                address2 = row[10]
                suburb = row[11]
                city = row[12]
                postal = row[13]
                email = row[16]
                status = row[17]
                dob = idnum
                gender = idnum
                date_created=date_created_old

                result=StatusData(
                    idnum=idnum,
                    cell=cell,
                    date_created=date_created,
                    salary=salary,
                    name=name,
                    surname=surname,
                    address1=address1,
                    address2=address2,
                    suburb=suburb,
                    city=city,
                    postal=postal,
                    email=email,
                    status=status,
                    dob=dob,
                    gender=gender
                )

                #append the dictionary
                rows.append(result.model_dump())
             
            #test the rows list if it has the right information
            
            #run queries by making updates on the information table,location table,contact table,employment table,car_table,finance table
        
        for i in range(5):

            if i==1:
                insert_list=get_status_tuple(rows,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")
                
                response=insert_data_into_finance_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"an error occurred while updating the finance table")

            elif i==2:
                insert_list=get_status_tuple(rows,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

                response=insert_data_into_location_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

            elif i==3:
                insert_list=get_status_tuple(rows,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

                response=insert_data_into_contact_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

            elif i==4:
                insert_list=get_status_tuple(rows,i)

                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

                response=insert_data_into_employment_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

            elif i==5:
                insert_list=get_status_tuple(rows,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

                response=insert_data_into_car_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

            else:
                insert_list=get_status_tuple(rows,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

                response=insert_data_into_finance_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"internal server error occurred")

        status_data_logger.info(f"{len(rows)} updated on finance, car , location, employment, and finance table(e)")
        return InsertStatusDataResponse(number_of_leads=leads_total,status=True,time_taken=total_time,information_table=f"{len(rows)} rows updated",contact_table=f"{len(rows)} rows updated",location_table=f"{len(rows)} updated",car_table=f"{len(rows)} rows updated",finance_table=f"{len(rows)} rows updated")
    
    except Exception as e:
        status_data_logger.error(f"{str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"{str(e)}")
    

@dedupe_routes.post("/insert_enriched_data",status_code=status.HTTP_200_OK,response_model=InsertEnrichedDataResponse)

async def insert_enriched_data(filename:str=Query(...,description="Provide the name for excel filename with status data")):
    try:
        data_extraction_time_start=time.time()

        try:
            all_sheets=pd.read_excel(filename + '.xlsx',sheet_name=None,dtype=str)
            sheets=all_sheets.keys()

        except Exception as e:
            #you have log this
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"{str(e)}")
        
        for sheet_name in sheets:
            sheet=pd.read_excel(filename,'.xlsx',sheet_name=sheet_name)
            sheet.to_csv(f"{sheet_name}.csv",index=False)
        
        enriched_data=[]

        detected_csv=[i for i in os.listdir() if i.endswith(".csv")]

        for d in detected_csv:
            csv_frame=pd.read_csv(d)
            csv_list=csv_frame.values().tolist()
            #append the contents of csv_list into the enriched data list
            enriched_data.extend(csv_list)
        
        total_leads=len(enriched_data)

        data_extraction_time_end=time.time()

        total_data_extraction_time=data_extraction_time_end - data_extraction_time_start

        rows_list=[]

        for e in enriched_data:

            Title=e[0]
            forename=e[1]
            lastname=e[2]
            IDNo=e[3]
            Race=e[4]
            gender=e[5]
            marital_status=e[6]
            line1=e[7]
            line2=e[8]
            line3=e[9]
            line4=e[10]
            PCode=e[11]
            Province=e[12]
            Home_Number=e[13]
            Work_Number=e[14]
            mobile_Number = e[15]
            mobile_Number2 = e[16]
            mobile_Number3 = e[17]
            mobile_Number4 = e[18]
            mobile_Number5 = e[19]
            mobile_Number6 = e[20]
            derived_income = e[21]

            cipro_reg = e[22]
            Deed_office_reg = e[23]
            vehicle_owner = e[24]
            cr_score_tu = e[25]
            monthly_expenditure = e[26]
            owns_cr_card = e[27]
            cr_card_rem_bal = e[28]
            owns_st_card = e[29]
            st_card_rem_bal = e[30]
            has_loan_acc = e[31]
            loan_acc_rem_bal = e[32]
            has_st_loan = e[33]
            st_loan_bal = e[34]
            has_1mth_loan = e[35]
            onemth_loan_bal = e[36]
            sti_insurance = e[37]
            has_sequestration = e[38]
            has_admin_order = e[39]
            under_debt_review = e[40]
            deceased_status = e[41]
            has_judgements = e[42]
            make = e[43]
            model = e[44]
            year = e[45]
            birth_date = e[3]

            new_convert = EnrichedData(Title=Title, forename=forename, lastname=lastname, IDNo=IDNo,Race=Race,gender=gender, Marital_Status=marital_status, line1=line1,line2=line2,line3=line3, line4=line4, PCode=PCode, Province=Province,Home_number=Home_Number, Work_number=Work_Number,mobile_Number=mobile_Number, mobile_Number2=mobile_Number2,mobile_Number3=mobile_Number3, mobile_Number4=mobile_Number4,mobile_Number5=mobile_Number5, mobile_Number6=mobile_Number6,derived_income=derived_income, cipro_reg=cipro_reg,Deed_office_reg=Deed_office_reg, vehicle_owner=vehicle_owner,cr_score_tu=cr_score_tu, monthly_expenditure=monthly_expenditure,owns_cr_card=owns_cr_card, cr_card_rem_bal=cr_card_rem_bal,owns_st_card=owns_st_card, st_card_rem_bal=st_card_rem_bal,has_loan_acc=has_loan_acc, loan_acc_rem_bal=loan_acc_rem_bal,has_st_loan=has_st_loan, st_loan_bal=st_loan_bal,has_1mth_loan=has_1mth_loan, onemth_loan_bal=onemth_loan_bal,sti_insurance=sti_insurance, has_sequestration=has_sequestration,has_admin_order=has_admin_order, under_debt_review=under_debt_review,deceased_status=deceased_status, has_judgements=has_judgements,make=make,model=model, year=year, birth_date = birth_date)

            rows_list.append(new_convert.model_dump())
        

        insertion_time1=time.time()

        for i in range(5):

            if i==1:
                insert_list=get_status_tuple(rows_list,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"an internal server error occurred while generating a list for the finance table")
                response=insert_data_into_finance_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An internal server error occurred while inserting data into the finance table")
            elif i==2:
                insert_list=get_status_tuple(rows_list,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while generating a list for location table")
                response=insert_data_into_location_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while inserting data on the location table")
            elif i==3:
                insert_list=get_status_tuple(rows_list,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while generating a list for the contact table")
                response=insert_data_into_contact_table(rows_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while inserting a list on the contact table")
            elif i==4:
                insert_list=get_status_tuple(rows_list,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while generating a list for the employment table")
                response=insert_data_into_employment_table(rows_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while insrting to the employment table")
            elif i==5:
                insert_list=get_status_tuple(rows_list,i)
                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while generating a list for the car table")
                response=insert_data_into_car_table(insert_list)

                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while inserting to the car table")
            else:
                insert_list=get_status_tuple(rows_list,i)

                if insert_list==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while generating a list for the finance table")
                response=insert_data_into_finance_table(insert_list)
                if response==False:
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error occurred while inserting data into the finance table")
                
        insertion_time2=time.time()
        
        total_insertion_time=insertion_time2 - insertion_time1

        return InsertEnrichedDataResponse(number_of_leads=total_leads,status=True,data_extraction_time=total_data_extraction_time,data_insertion_time=total_insertion_time,contact_table=f"Contact table updated with:{len(enriched_data)} records",location_table=f"location table updated with:{len(enriched_data)} records",car_table=f"car table updated with:{len(enriched_data)} records",finance_table=f"finance table updated with:{len(enriched_data)} recorda")
    except Exception as e:
        #you need log here
        status_data_logger.error(f"{str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"{str(e)}")

@dedupe_routes.post("/add_manual_dedupe_list2",status_code=status.HTTP_201_CREATED)

async def add_manual_dedupe_list2(filename:Annotated[UploadFile,File()],camp_code:str=Query(description="Vicidial Campaign Code"),session:Session=Depends(get_session)):
   
    try:
        #guide against reading the wrong file
        try:
            file_content=await filename.read()

        except Exception as e:
            dedupe_logger.info(f"failed to read the uploaded file content:{e}")
            dedupe_logger.error(e)
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"failed to read the uploaded file content")
        try:
            wb=openpyxl.load_workbook(BytesIO(file_content))
        
        except Exception as e:
            dedupe_logger.info("Error occurred while reading from the workbook")
            dedupe_logger.error(e)
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,detail=f"the file uploaded is not a valid or readable Excel workbook (.xlsx)")
        
        rows=[]

        for sheet in wb.worksheets:

            for row_tuple in sheet.iter_rows(min_row=1,max_col=2,values_only=True):
                id_num,cell_num=row_tuple[0],row_tuple[1]
                id_num_str=str(id_num) if id_num is not None else None
                line=[id_num_str,cell_num]
                rows.append(line)
        
        suffix=''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(4))

        key=filename.filename + suffix
        #data to insert in the campaign_dedupe table

        data=[(r[0],r[1],camp_code,'P',key) for r in rows]

        #dedupe_models=[]

        campaigns=[Campaign_Dedupe(id=id,cell=cell,campaign_name=camp_code,status=status,key=key) for id,cell,camp_code,status,key in data]

        try:
            session.add_all(campaigns)
            session.commit()
            session.close()
            dedupe_logger.info(f"inserted {len(campaigns)} records into the campaign dedupe table")

        except Exception as e:
            session.rollback()
            session.close()
            dedupe_logger.exception(f"An exception occurred while:{e}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"An exception occurred while inserting data in the campaign dedupe table")



        # for item in data:
        #     model_instance=Campaign_Dedupe(id=item[0],cell_number=item[1],campaign_name=item[2],status=item[3],code=item[4])
        #     dedupe_models.append(model_instance)
        
        dedupe_logger.info("connecting to the database for manual dedupe upload")
       
        update_data=[(r[1],'DEDUPE') for r in rows]

        query="""
                INSERT INTO info_tbl(cell,extra_info)
                VALUES(:cell,:extra_info)
                ON CONFLICT(cell)
                DO UPDATE SET extra_info=EXCLUDED.extra_info
                WHERE info_tbl.cell = EXCLUDED.cell
             """
        
        try:
            # insert and perform update 
            with engine.begin() as conn:
                conn.execute(text(query),[{"cell":cell,"extra_info":info} for cell,info in update_data])
                conn.close()

            dedupe_logger.info(f"inserted records of length:{len(update_data)} on the info_tbl(information table)")

        except Exception as e:
            conn.rollback()
            dedupe_logger.exception(f"an exception occurred while updating info_tbl with:{len(update_data)} data:{e}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"an internal server error occurred while inserting data into the info_tbl")
        
        
        # with engine.begin() as conn:

        #     conn.execute(text(query),[{"cell":cell,"extra_info":info} for cell,info in update_data])


        # upsert_data=[{"cell":item[0],"extra_info":item[1]} for item in update_data]

        # insert_statement=insert(info_tbl).values(upsert_data)

        # #potential hazards
        # upsert_statement=insert_statement.on_conflict_do_update(index_elements=['cell'],set={"extra_info":insert_statement.excluded.extra_info})
        
        # try:
        #     session.add_all(dedupe_models)
        #     dedupe_logger.info(f"dedupe batches successfully added:{len(dedupe_models)}")
        #     session.exec(upsert_statement)
        #     session.commit()
        #     dedupe_logger.info("Successfully updated number and extra information cell")
        # except Exception as e:
        #     dedupe_logger.error(e)
        #     session.rollback()
        #     raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail=f"error occuured while processing the data")
        
        return ManualDedupeListReturn(Success=True,Information_Table=f"{len(update_data)} records updated on the information table",Campaign_Dedupe_Table=f"{len(campaigns)} inserted on the campaign dedupe table",Key=key)
       

    except Exception as e:
        dedupe_logger.critical(e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,detail="An internal server error")
    

